from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.utils import timezone
from django.utils.timesince import timesince
from django.views.decorators.http import require_POST  # <-- ADD THIS LINE
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.cache import never_cache
from datetime import timedelta, datetime
from django.http import JsonResponse, HttpResponse
from functools import wraps
from .models import *
from .forms import UserUpdateForm, ProfileUpdateForm


def login_required_json(view_func):
    """For AJAX requests, return 401 JSON instead of redirecting to login."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if request.user.is_authenticated:
            return view_func(request, *args, **kwargs)
        # Prefer JSON response for AJAX so frontend can show "log in" message
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.META.get('HTTP_ACCEPT', ''):
            return JsonResponse({'error': 'Login required', 'products': []}, status=401)
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(request.get_full_path())
    return _wrapped
from .utils import *
from .notifications import NotificationService, check_low_stock_and_notify  # If you created this
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference, LineChart
import csv

# ========== AUTHENTICATION VIEWS ==========

@never_cache
@csrf_protect
def login_view(request):
    # If user is already authenticated, redirect to dashboard
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        next_url = request.POST.get('next', request.GET.get('next', 'dashboard'))
        
        # Validate input
        if not username or not password:
            messages.error(request, 'Please provide both username and password.')
            return render(request, 'registration/login.html')
        
        # Authenticate user
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                messages.success(request, f'Login Successfully!')
                
                # Check if "Remember me" is checked
                if not request.POST.get('remember_me'):
                    # Session expires on browser close
                    request.session.set_expiry(0)
                    
                # Check if next_url is safe
                if next_url and next_url != 'None' and next_url.startswith('/'):
                    return redirect(next_url)
                return redirect('dashboard')
            else:
                messages.error(request, 'Your account is disabled.')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'registration/login.html')
@never_cache
@login_required
def logout_view(request):
    # Explicitly clear auth/session state and force a new anonymous session.
    logout(request)
    request.session.flush()

    response = redirect('login')
    response.delete_cookie('sessionid')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@never_cache
@login_required
def user_profile(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        form_type = request.POST.get('form_type', 'profile')

        if form_type == 'password':
            from django.contrib.auth import update_session_auth_hash
            current = request.POST.get('current_password')
            new1 = request.POST.get('new_password1')
            new2 = request.POST.get('new_password2')
            if not request.user.check_password(current):
                messages.error(request, 'Current password is incorrect.')
            elif not new1 or len(new1) < 6:
                messages.error(request, 'New password must be at least 6 characters.')
            elif new1 != new2:
                messages.error(request, 'New passwords do not match.')
            else:
                request.user.set_password(new1)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Password changed successfully.')
            return redirect('user_profile')

        # Profile form
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Your profile has been updated.')
            return redirect('user_profile')
        messages.error(request, 'Please correct the errors below.')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=profile)

    return render(
        request,
        'core/profile.html',
        {
            'profile': profile,
            'user_form': user_form,
            'profile_form': profile_form,
            'role_label': profile.get_role_display(),
        },
    )

def is_admin_or_manager(user):
    return user.is_superuser or (hasattr(user, 'profile') and user.profile.role in ['admin', 'production_admin', 'manager'])


def is_production_team(user):
    return hasattr(user, 'profile') and user.profile.role == 'production_admin'

# ========== DASHBOARD VIEWS ==========

@login_required
def dashboard(request):
    today = timezone.localdate()
    start_of_month = today.replace(day=1)
    yesterday = today - timedelta(days=1)
    start_30 = today - timedelta(days=30)
    tz = timezone.get_current_timezone()

    # --- Consolidate sales queries into one pass ---
    from django.db.models import Case, When, DecimalField
    sales_agg = Order.objects.aggregate(
        daily=Sum(Case(When(created_at__date=today, then='total'), output_field=DecimalField())),
        yesterday=Sum(Case(When(created_at__date=yesterday, then='total'), output_field=DecimalField())),
        monthly=Sum(Case(When(created_at__date__gte=start_of_month, then='total'), output_field=DecimalField())),
        total_today=Count(Case(When(created_at__date=today, then='id'))),
        pending=Count(Case(When(status='pending', then='id'))),
    )
    daily_sales   = sales_agg['daily'] or 0
    yesterday_sales = sales_agg['yesterday'] or 0
    monthly_sales = sales_agg['monthly'] or 0
    total_orders  = sales_agg['total_today'] or 0
    pending_orders = sales_agg['pending'] or 0

    if yesterday_sales > 0:
        sales_change_pct = ((daily_sales - yesterday_sales) / yesterday_sales) * 100
    elif daily_sales > 0:
        sales_change_pct = 100
    else:
        sales_change_pct = 0

    # Low stock
    low_stock_products = list(Product.objects.filter(
        stock__lte=F('low_stock_threshold'), is_archived=False
    ).only('id', 'name', 'stock', 'low_stock_threshold')[:10])

    # Best sellers + category + top products — all from one queryset scan
    oi_base = OrderItem.objects.filter(order__created_at__date__gte=start_30)

    best_sellers = list(oi_base.values('product__name', 'product__id').annotate(
        quantity=Sum('quantity'), total_sales=Sum('subtotal')
    ).order_by('-quantity')[:5])

    if not best_sellers:
        best_sellers = list(
            Order.objects.exclude(product_name__isnull=True).exclude(product_name='')
            .values('product_name').annotate(quantity=Count('id'), total_sales=Sum('total'))
            .order_by('-total_sales')[:5]
        )
        for b in best_sellers:
            b['product__name'] = b.pop('product_name', '')

    sales_by_category_oi = list(oi_base.values('product__category').annotate(
        total_sales=Sum('subtotal'), total_qty=Sum('quantity')
    ).order_by('-total_sales'))

    top_products_oi = list(oi_base.values('product__name').annotate(
        total_sales=Sum('subtotal'), total_qty=Sum('quantity')
    ).order_by('-total_sales')[:10])

    # Build chart data
    category_labels, category_sales_data, category_qty = [], [], []
    if sales_by_category_oi:
        cat_map = dict(Product.CATEGORY_CHOICES)
        for item in sales_by_category_oi:
            cat = cat_map.get(item['product__category'], str(item['product__category']).title())
            category_labels.append(cat)
            category_sales_data.append(float(item['total_sales']))
            category_qty.append(int(item['total_qty'] or 0))

    if top_products_oi:
        top_product_labels = [p.get('product__name', '')[:20] for p in top_products_oi]
        top_product_sales  = [float(p['total_sales']) for p in top_products_oi]
        top_product_qty    = [p.get('total_qty', 0) for p in top_products_oi]
    else:
        top_product_labels, top_product_sales, top_product_qty = [], [], []

    # Sales graph — last 7 days
    sales_qs = (
        Order.objects.filter(created_at__date__gte=today - timedelta(days=6))
        .annotate(day=TruncDay('created_at', tzinfo=tz))
        .values('day').annotate(total=Sum('total')).order_by('day')
    )
    sales_by_day = {row['day'].date(): float(row['total'] or 0) for row in sales_qs}
    last_7_days, sales_data = [], []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        last_7_days.append(d.strftime('%Y-%m-%d'))
        sales_data.append(sales_by_day.get(d, 0))

    # Recent orders
    recent_orders = Order.objects.select_related('customer').prefetch_related('items').order_by('-created_at')[:5]

    # Products + customers for New Order modal
    products  = Product.objects.filter(is_archived=False).only('id', 'name', 'price', 'stock', 'code').order_by('name')[:100]
    customers = Customer.objects.only('id', 'name').order_by('name')[:50]

    context = {
        'daily_sales': daily_sales,
        'monthly_sales': monthly_sales,
        'sales_change_pct': round(sales_change_pct, 1),
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'low_stock_products': low_stock_products,
        'best_sellers': best_sellers,
        'recent_orders': recent_orders,
        'sales_labels': json.dumps(last_7_days),
        'sales_data': json.dumps(sales_data),
        'category_labels': json.dumps(category_labels),
        'category_sales': json.dumps(category_sales_data),
        'category_qty': json.dumps(category_qty),
        'top_product_labels': json.dumps(top_product_labels),
        'top_product_sales': json.dumps(top_product_sales),
        'top_product_qty': json.dumps(top_product_qty),
        'products': products,
        'customers': customers,
    }

    return render(request, 'core/dashboard.html', context)

# ========== PRODUCT VIEWS ==========

@login_required
def product_list(request):
    products = Product.objects.filter(is_archived=False).only(
        'id', 'code', 'name', 'category', 'price', 'cost', 'stock',
        'low_stock_threshold', 'image', 'is_available', 'is_new_arrival', 'is_best_seller'
    )
    archived_products = Product.objects.filter(is_archived=True).only(
        'id', 'code', 'name', 'category', 'archived_at'
    ).order_by('-archived_at', '-updated_at')[:50]
    categories = Product.CATEGORY_CHOICES
    raw_materials = RawMaterial.objects.only('id', 'name', 'unit').order_by('name')

    category = request.GET.get('category')
    if category:
        products = products.filter(category=category)

    search = request.GET.get('search')
    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )

    context = {
        'products': products,
        'archived_products': archived_products,
        'categories': categories,
        'raw_materials': raw_materials,
    }
    return render(request, 'core/product_list.html', context)

# ========== INVENTORY VIEWS ==========

@login_required
def inventory_status(request):
    products = Product.objects.filter(is_archived=False).order_by('-stock').only(
        'id', 'code', 'name', 'category', 'price', 'cost', 'stock',
        'low_stock_threshold', 'image', 'is_available'
    )
    raw_materials = RawMaterial.objects.all().only(
        'id', 'name', 'unit', 'stock_quantity', 'low_stock_threshold',
        'reorder_point', 'supplier', 'cost_per_unit'
    )

    # Calculate inventory values in Python (avoid extra annotation query)
    products = list(products)
    for product in products:
        product.inventory_value = product.stock * product.price

    raw_materials = list(raw_materials)
    for material in raw_materials:
        material.inventory_value = material.stock_quantity * material.cost_per_unit

    # Low stock items (filter from already-fetched lists)
    low_stock_products = [p for p in products if p.stock <= p.low_stock_threshold]
    low_stock_materials = [m for m in raw_materials if m.stock_quantity <= m.low_stock_threshold]

    # Recent transactions
    recent_transactions = InventoryTransaction.objects.select_related(
        'product', 'created_by'
    ).order_by('-created_at')[:20]

    context = {
        'products': products,
        'raw_materials': raw_materials,
        'low_stock_products': low_stock_products,
        'low_stock_materials': low_stock_materials,
        'recent_transactions': recent_transactions,
    }
    return render(request, 'core/inventory_status.html', context)

# ========== ORDER VIEWS ==========

@login_required
def order_list(request):
    orders = Order.objects.all().select_related('customer').prefetch_related('items')

    # Production team only sees incoming/in-process orders
    if is_production_team(request.user):
        orders = orders.filter(status__in=['pending', 'confirmed', 'preparing', 'ready'])

    # Filter by status
    status = request.GET.get('status')
    if status:
        orders = orders.filter(status=status)

    # Filter by date
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    # Limit to 200 most recent to keep page fast
    orders = orders.order_by('-created_at')[:200]

    context = {
        'orders': orders,
        'status_choices': Order.ORDER_STATUS,
    }
    return render(request, 'core/order_list.html', context)

@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    # Safely get customer — handles UUID customer_id from mobile app
    customer = None
    try:
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT customer_id FROM core_order WHERE id = %s", [pk])
            row = cursor.fetchone()
            if row and row[0]:
                try:
                    cid = int(row[0])
                    customer = Customer.objects.filter(pk=cid).first()
                except (ValueError, TypeError):
                    customer = None
    except Exception:
        customer = None
    context = {
        'order': order,
        'customer': customer,
    }
    return render(request, 'core/order_detail.html', context)

@login_required
def confirm_order(request, pk):
    if request.method == 'POST':
        order = get_object_or_404(Order, pk=pk)
        
        if order.status == 'pending':
            # Check stock availability
            insufficient_stock = []
            for item in order.items.all():
                if item.product.stock < item.quantity:
                    insufficient_stock.append(f"{item.product.name} (available: {item.product.stock}, needed: {item.quantity})")
            
            if insufficient_stock:
                messages.error(request, f"Insufficient stock for: {', '.join(insufficient_stock)}")
                return redirect('order_detail', pk=pk)
            
            try:
                with transaction.atomic():
                    for item in order.items.all():
                        product = Product.objects.select_for_update().get(pk=item.product.pk)
                        old_stock = product.stock
                        product.stock -= item.quantity
                        product.save()

                        InventoryTransaction.objects.create(
                            product=product,
                            transaction_type='out',
                            quantity=item.quantity,
                            previous_stock=old_stock,
                            new_stock=product.stock,
                            reference=order.order_number,
                            notes="Order confirmation",
                            created_by=request.user
                        )

                        # Notify if low stock after deduction
                        from .notifications import notify_if_low_stock, notify_if_material_low
                        notify_if_low_stock(product)

                        for recipe in product.recipe.all():
                            material = recipe.raw_material
                            needed_qty = recipe.quantity * item.quantity
                            if material.stock_quantity >= needed_qty:
                                old_material_stock = material.stock_quantity
                                material.stock_quantity -= needed_qty
                                material.save()
                                RawMaterialTransaction.objects.create(
                                    raw_material=material,
                                    transaction_type='out',
                                    quantity=needed_qty,
                                    previous_stock=old_material_stock,
                                    new_stock=material.stock_quantity,
                                    reference=order.order_number,
                                    notes=f"Used for {product.name}",
                                    created_by=request.user
                                )
                                notify_if_material_low(material)

                    # Use raw update to avoid loading UUID customer FK
                    Order.objects.filter(pk=order.pk).update(status='confirmed')
                
                # Notification — safely handle UUID customer_id
                try:
                    from django.db import connection
                    from .notifications import NotificationService
                    # Notify admins about confirmed order
                    NotificationService.notify_admins(
                        title=f"Order #{order.order_number} Confirmed",
                        message=f"Order #{order.order_number} has been confirmed and is being prepared.",
                        notification_type='order',
                        priority='medium',
                        link=f"/orders/{order.pk}/",
                        action_text='View Order'
                    )
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT customer_id FROM core_order WHERE id = %s", [order.pk])
                        row = cursor.fetchone()
                        if row and row[0]:
                            try:
                                cid = int(row[0])
                                customer = Customer.objects.filter(pk=cid).first()
                                if customer and customer.user:
                                    Notification.objects.create(
                                        title=f"Order #{order.order_number} Confirmed",
                                        message="Your order has been confirmed and is being prepared.",
                                        notification_type='order',
                                        recipient_type='customer',
                                        recipient_user=customer.user
                                    )
                            except (ValueError, TypeError):
                                pass
                except Exception:
                    pass
                
                messages.success(request, f'Order #{order.order_number} confirmed successfully.')
            
            except Exception as e:
                messages.error(request, f'Error confirming order: {str(e)}')
    
    return redirect('order_detail', pk=pk)


@login_required
@require_POST
def update_order_status(request, pk):
    """Update order status (confirmed → preparing → ready → delivered, or cancel)"""
    order = get_object_or_404(Order, pk=pk)
    new_status = request.POST.get('status')

    valid_statuses = dict(Order.ORDER_STATUS).keys()
    if new_status not in valid_statuses:
        messages.error(request, 'Invalid status.')
        return redirect('order_detail', pk=pk)

    # Prevent going backwards (except cancellation)
    status_flow = ['pending', 'confirmed', 'preparing', 'ready', 'out_for_delivery', 'delivered']
    if new_status != 'cancelled' and new_status in status_flow and order.status in status_flow:
        current_idx = status_flow.index(order.status)
        new_idx = status_flow.index(new_status)
        if new_idx < current_idx:
            messages.error(request, f'Cannot move order back from "{order.get_status_display()}" to "{dict(Order.ORDER_STATUS)[new_status]}".')
            return redirect('order_detail', pk=pk)

    old_status = order.status
    Order.objects.filter(pk=pk).update(status=new_status)

    # Notify admins and customer
    try:
        from .notifications import NotificationService
        status_label = dict(Order.ORDER_STATUS).get(new_status, new_status)
        NotificationService.notify_admins(
            title=f"Order #{order.order_number} → {status_label}",
            message=f"Order status updated from {dict(Order.ORDER_STATUS).get(old_status, old_status)} to {status_label}.",
            notification_type='order',
            priority='medium',
            link=f"/orders/{order.pk}/",
            action_text='View Order'
        )
        # Notify customer if linked
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT customer_id FROM core_order WHERE id = %s", [pk])
            row = cursor.fetchone()
            if row and row[0]:
                try:
                    cid = int(row[0])
                    customer = Customer.objects.filter(pk=cid).first()
                    if customer and customer.user:
                        Notification.objects.create(
                            title=f"Order #{order.order_number}: {status_label}",
                            message=f"Your order status is now: {status_label}.",
                            notification_type='order',
                            recipient_type='customer',
                            recipient_user=customer.user,
                            link=f"/orders/{order.pk}/"
                        )
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass

    messages.success(request, f'Order #{order.order_number} status updated to "{dict(Order.ORDER_STATUS).get(new_status, new_status)}".')
    return redirect('order_detail', pk=pk)


@login_required
@require_POST
def update_payment_status(request, pk):
    """Update payment status for an order"""
    order = get_object_or_404(Order, pk=pk)
    new_payment_status = request.POST.get('payment_status')

    valid_payment_statuses = dict(Order.PAYMENT_STATUS).keys()
    if new_payment_status not in valid_payment_statuses:
        messages.error(request, 'Invalid payment status.')
        return redirect('order_detail', pk=pk)

    Order.objects.filter(pk=pk).update(payment_status=new_payment_status)
    messages.success(request, f'Payment status updated to "{dict(Order.PAYMENT_STATUS).get(new_payment_status, new_payment_status)}".')
    return redirect('order_detail', pk=pk)

# ========== SALES VIEWS ==========

@login_required
def sales_report(request):
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = (timezone.now() - timedelta(days=30)).date()
    else:
        try:
            date_from = datetime.strptime(str(date_from), '%Y-%m-%d').date()
        except ValueError:
            date_from = (timezone.now() - timedelta(days=30)).date()
    if not date_to:
        date_to = timezone.now().date()
    else:
        try:
            date_to = datetime.strptime(str(date_to), '%Y-%m-%d').date()
        except ValueError:
            date_to = timezone.now().date()
    
    # Include all orders in the selected date range (regardless of payment status)
    orders = Order.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    )
    
    # Summary statistics
    total_sales = orders.aggregate(total=Sum('total'))['total'] or 0
    total_orders = orders.count()
    average_order = total_sales / total_orders if total_orders > 0 else 0
    
    # Sales by category — from OrderItem, fallback to product_name in orders
    sales_by_category = list(OrderItem.objects.filter(
        order__in=orders
    ).values('product__category').annotate(
        total=Sum('subtotal'),
        quantity=Sum('quantity')
    ).order_by('-total'))

    # Fallback: group by product category using product_name matched to Product table
    if not sales_by_category:
        from django.db.models import FloatField
        # Get product_name from orders and match to products to get category
        order_products = list(
            orders.exclude(product_name__isnull=True).exclude(product_name='')
            .values('product_name')
            .annotate(total=Sum('total'), quantity=Count('id'))
        )
        # Match product names to categories
        cat_totals = {}
        for op in order_products:
            product = Product.objects.filter(name__iexact=op['product_name']).first()
            if product:
                cat = product.get_category_display()
            else:
                cat = 'Other'
            if cat not in cat_totals:
                cat_totals[cat] = {'total': 0, 'quantity': 0, 'product__category': cat}
            cat_totals[cat]['total'] += float(op['total'])
            cat_totals[cat]['quantity'] += op['quantity']
        sales_by_category = sorted(cat_totals.values(), key=lambda x: -x['total'])

    # Top products — from OrderItem, fallback to product_name in orders
    top_products = list(OrderItem.objects.filter(
        order__in=orders
    ).values('product__name').annotate(
        quantity=Sum('quantity'),
        revenue=Sum('subtotal')
    ).order_by('-revenue')[:10])

    # Fallback: use product_name stored in orders if no OrderItems
    if not top_products:
        top_products = list(
            orders.exclude(product_name__isnull=True).exclude(product_name='')
            .values('product_name')
            .annotate(quantity=Count('id'), revenue=Sum('total'))
            .order_by('-revenue')[:10]
        )
        for p in top_products:
            p['product__name'] = p.pop('product_name', '')

    # Daily sales for graph — use Manila timezone for accurate date grouping
    tz = timezone.get_current_timezone()
    daily_sales = orders.annotate(
        day=TruncDay('created_at', tzinfo=tz)
    ).values('day').annotate(
        total=Sum('total')
    ).order_by('day')

    sales_dates = [item['day'].astimezone(tz).strftime('%Y-%m-%d') for item in daily_sales]
    sales_amounts = [float(item['total']) for item in daily_sales]
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'total_sales': total_sales,
        'total_orders': total_orders,
        'average_order': average_order,
        'sales_by_category': sales_by_category,
        'sales_dates': json.dumps(sales_dates),
        'sales_amounts': json.dumps(sales_amounts),
        'top_products': top_products,
    }
    return render(request, 'core/sales_report.html', context)


@login_required
def production_cost_analytics(request):
    """Production cost analytics and profit margin analysis."""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if not date_from:
        date_from = (timezone.now() - timedelta(days=30)).date()
    else:
        try:
            date_from = datetime.strptime(str(date_from), '%Y-%m-%d').date()
        except ValueError:
            date_from = (timezone.now() - timedelta(days=30)).date()
    if not date_to:
        date_to = timezone.now().date()
    else:
        try:
            date_to = datetime.strptime(str(date_to), '%Y-%m-%d').date()
        except ValueError:
            date_to = timezone.now().date()

    analysis = calculate_profit_analysis(start_date=date_from, end_date=date_to)
    analysis['date_from'] = date_from
    analysis['date_to'] = date_to

    # Product-level margins
    product_margins = []
    for name, data in analysis.get('products', {}).items():
        product_margins.append({
            'name': name,
            'revenue': data['revenue'],
            'cost': data['cost'],
            'profit': data['profit'],
            'margin': data['margin'],
            'quantity': data['quantity'],
        })
    product_margins.sort(key=lambda x: -x['revenue'])

    context = {
        'analysis': analysis,
        'product_margins': product_margins,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'core/production_analytics.html', context)


@login_required
def export_sales_report(request):
    # Get date range from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = (timezone.now() - timedelta(days=30)).date()
    else:
        try:
            date_from = datetime.strptime(str(date_from), '%Y-%m-%d').date()
        except ValueError:
            date_from = (timezone.now() - timedelta(days=30)).date()
    if not date_to:
        date_to = timezone.now().date()
    else:
        try:
            date_to = datetime.strptime(str(date_to), '%Y-%m-%d').date()
        except ValueError:
            date_to = timezone.now().date()
    
    # Create Excel workbook
    wb = Workbook()
    
    # Sales Summary Sheet
    ws1 = wb.active
    ws1.title = "Sales Summary"
    
    # Headers
    headers = ['Date', 'Order Number', 'Customer', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total', 'Payment Status']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Get orders
    orders = Order.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).order_by('-created_at')
    
    # Write data
    row = 2
    for order in orders:
        items_count = order.items.count()
        ws1.cell(row=row, column=1, value=order.created_at.strftime('%Y-%m-%d'))
        ws1.cell(row=row, column=2, value=order.order_number)
        ws1.cell(row=row, column=3, value=order.customer.name if order.customer else 'Walk-in')
        ws1.cell(row=row, column=4, value=items_count)
        ws1.cell(row=row, column=5, value=float(order.subtotal))
        ws1.cell(row=row, column=6, value=float(order.tax))
        ws1.cell(row=row, column=7, value=float(order.discount))
        ws1.cell(row=row, column=8, value=float(order.total))
        ws1.cell(row=row, column=9, value=order.payment_status)
        row += 1
    
    # Add totals row
    if row > 2:
        totals_row = row + 1
        ws1.cell(row=totals_row, column=5, value=f"=SUM(E2:E{row-1})")
        ws1.cell(row=totals_row, column=8, value=f"=SUM(H2:H{row-1})")
        ws1.cell(row=totals_row, column=5).font = Font(bold=True)
        ws1.cell(row=totals_row, column=8).font = Font(bold=True)
    
    # Create chart
    chart = LineChart()
    chart.title = "Daily Sales Trend"
    chart.style = 13
    chart.y_axis.title = 'Sales Amount'
    chart.x_axis.title = 'Date'
    
    # Only add chart when there is at least one data row.
    if row > 2:
        data = Reference(ws1, min_col=8, min_row=1, max_row=row-1, max_col=8)
        dates = Reference(ws1, min_col=1, min_row=2, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)
        ws1.add_chart(chart, "K2")
    
    # Products Sheet
    ws2 = wb.create_sheet("Top Products")
    
    headers = ['Product', 'Category', 'Quantity Sold', 'Revenue']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    top_products = OrderItem.objects.filter(
        order__in=orders
    ).values(
        'product__name', 'product__category'
    ).annotate(
        quantity=Sum('quantity'),
        revenue=Sum('subtotal')
    ).order_by('-revenue')[:20]
    
    row = 2
    for item in top_products:
        ws2.cell(row=row, column=1, value=item['product__name'])
        ws2.cell(row=row, column=2, value=item['product__category'])
        ws2.cell(row=row, column=3, value=item['quantity'])
        ws2.cell(row=row, column=4, value=float(item['revenue']))
        row += 1
    
    # Set column widths
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{date_from}_to_{date_to}.xlsx"'
    
    wb.save(response)
    return response

# ========== FORECAST VIEWS ==========

@login_required
def forecast(request):
    # Include all products so forecasting still works even if stock is 0
    products = Product.objects.all().order_by('name')
    
    # Get forecasts
    forecasts = SalesForecast.objects.filter(
        forecast_date__gte=timezone.now().date()
    ).select_related('product').order_by('forecast_date')[:30]
    
    # Prepare chart data
    forecast_dates = []
    forecast_values = []
    for forecast in forecasts:
        forecast_dates.append(forecast.forecast_date.strftime('%Y-%m-%d'))
        forecast_values.append(forecast.predicted_quantity)
    
    # Get historical data for comparison (last 30 days)
    historical_data = OrderItem.objects.filter(
        order__created_at__date__gte=timezone.now().date() - timedelta(days=30),
        order__created_at__date__lte=timezone.now().date(),
    ).values(
        'order__created_at__date'
    ).annotate(
        total=Sum('quantity')
    ).order_by('order__created_at__date')
    
    historical_dates = [item['order__created_at__date'].strftime('%Y-%m-%d') for item in historical_data]
    historical_values = [item['total'] for item in historical_data]
    
    context = {
        'products': products,
        'forecasts': forecasts,
        'forecast_dates': json.dumps(forecast_dates),
        'forecast_values': json.dumps(forecast_values),
        'historical_dates': json.dumps(historical_dates),
        'historical_values': json.dumps(historical_values),
    }
    return render(request, 'core/forecast.html', context)

@login_required
def run_forecast(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        days = min(int(request.POST.get('days', 30)), 90)  # cap at 90

        created_count = 0
        today = timezone.now().date()
        try:
            if product_id:
                product = get_object_or_404(Product, id=product_id)
                # Clear stale forecasts for this product before regenerating
                SalesForecast.objects.filter(product=product, forecast_date__lt=today).delete()
                result = generate_simple_forecast(product, days)
                created_count = len(result or [])
            else:
                # Clear all stale past forecasts
                SalesForecast.objects.filter(forecast_date__lt=today).delete()
                # Run for all active products — bulk_create makes this fast
                products = Product.objects.filter(is_archived=False).order_by('name')
                for product in products:
                    try:
                        result = generate_simple_forecast(product, days)
                        created_count += len(result or [])
                    except Exception:
                        continue
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'Forecast error: {str(e)}')
            return redirect('forecast')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'count': created_count})
        messages.success(request, f'Forecast generated! ({created_count} entries)')
        return redirect('forecast')

    return redirect('forecast')


@login_required
def forecast_data(request):
    """JSON endpoint — returns forecast + historical data for the chart."""
    days = int(request.GET.get('days', 30))
    product_id = request.GET.get('product_id')

    today = timezone.now().date()

    if product_id:
        # Single product — show its own N-day series
        qs = SalesForecast.objects.filter(
            forecast_date__gte=today,
            product_id=product_id,
        ).select_related('product').order_by('forecast_date')[:days]

        forecasts_data = []
        forecast_dates = []
        forecast_values = []
        for f in qs:
            forecast_dates.append(f.forecast_date.strftime('%Y-%m-%d'))
            forecast_values.append(f.predicted_quantity)
            forecasts_data.append({
                'product_name': f.product.name,
                'forecast_date': f.forecast_date.strftime('%Y-%m-%d'),
                'predicted_quantity': f.predicted_quantity,
                'confidence_lower': f.confidence_lower,
                'confidence_upper': f.confidence_upper,
                'model_used': f.model_used,
            })

        hist_qs = OrderItem.objects.filter(
            product_id=product_id,
            order__created_at__date__gte=today - timedelta(days=days),
            order__created_at__date__lte=today,
        ).values('order__created_at__date').annotate(
            total=Sum('quantity')
        ).order_by('order__created_at__date')

    else:
        # All products — aggregate by date so each date = 1 data point on the chart
        agg_qs = SalesForecast.objects.filter(
            forecast_date__gte=today,
        ).values('forecast_date').annotate(
            total_qty=Sum('predicted_quantity'),
            total_lower=Sum('confidence_lower'),
            total_upper=Sum('confidence_upper'),
        ).order_by('forecast_date')[:days]

        forecasts_data = []
        forecast_dates = []
        forecast_values = []
        for row in agg_qs:
            d = row['forecast_date'].strftime('%Y-%m-%d')
            forecast_dates.append(d)
            forecast_values.append(row['total_qty'])
            forecasts_data.append({
                'product_name': 'All Products',
                'forecast_date': d,
                'predicted_quantity': row['total_qty'],
                'confidence_lower': row['total_lower'],
                'confidence_upper': row['total_upper'],
                'model_used': 'WMA',
            })

        hist_qs = OrderItem.objects.filter(
            order__created_at__date__gte=today - timedelta(days=days),
            order__created_at__date__lte=today,
        ).values('order__created_at__date').annotate(
            total=Sum('quantity')
        ).order_by('order__created_at__date')

    historical_dates  = [i['order__created_at__date'].strftime('%Y-%m-%d') for i in hist_qs]
    historical_values = [i['total'] for i in hist_qs]

    return JsonResponse({
        'forecasts': forecasts_data,
        'forecast_dates': forecast_dates,
        'forecast_values': forecast_values,
        'historical_dates': historical_dates,
        'historical_values': historical_values,
    })

# ========== IMPORT/EXPORT VIEWS ==========

def _process_import_rows(import_type, df, file_name, request, errors_list):
    """Process import rows - returns (success_count, error_count, error_rows for report)"""
    success_count = 0
    error_count = 0
    error_rows = []  # List of dicts for error report: {row_num, data, error}
    
    if import_type == 'product':
        required = ['code', 'name', 'category', 'price']
        for index, row in df.iterrows():
            row_num = index + 2  # 1-based, +1 for header
            row_data = row.to_dict()
            try:
                code = row.get('code')
                if pd.isna(code) or not str(code).strip():
                    raise ValueError("Code is required")
                category = str(row.get('category', '')).strip().lower()
                valid_cats = ['bread', 'pastries', 'croissants', 'beverages', 'spread']
                if category not in valid_cats:
                    raise ValueError(f"Invalid category. Use: {', '.join(valid_cats)}")
                Product.objects.update_or_create(
                    code=str(code).strip(),
                    defaults={
                        'name': str(row.get('name', '')).strip() or 'Unknown',
                        'category': category,
                        'price': float(row.get('price', 0)),
                        'cost': float(row.get('cost', 0)),
                        'stock': int(row.get('stock', 0)),
                        'description': str(row.get('description', '')),
                    }
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                error_rows.append({'row': row_num, 'data': row_data, 'error': str(e)})
                errors_list.append(f"Row {row_num}: {str(e)}")
    
    elif import_type == 'customer':
        for index, row in df.iterrows():
            row_num = index + 2
            row_data = row.to_dict()
            try:
                email = row.get('email')
                if pd.isna(email) or not str(email).strip():
                    raise ValueError("Email is required")
                Customer.objects.update_or_create(
                    email=str(email).strip(),
                    defaults={
                        'name': str(row.get('name', '')).strip() or 'Unknown',
                        'phone': str(row.get('phone', '')),
                        'address': str(row.get('address', '')),
                    }
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                error_rows.append({'row': row_num, 'data': row_data, 'error': str(e)})
                errors_list.append(f"Row {row_num}: {str(e)}")
    
    elif import_type == 'inventory':
        for index, row in df.iterrows():
            row_num = index + 2
            row_data = row.to_dict()
            try:
                product = Product.objects.get(code=str(row.get('product_code', '')).strip())
                new_stock = int(row.get('new_stock', 0))
                old_stock = product.stock
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='adjustment',
                    quantity=new_stock - old_stock,
                    previous_stock=old_stock,
                    new_stock=new_stock,
                    notes=f"Bulk import from {file_name}",
                    created_by=request.user
                )
                product.stock = new_stock
                product.save()
                success_count += 1
            except Exception as e:
                error_count += 1
                error_rows.append({'row': row_num, 'data': row_data, 'error': str(e)})
                errors_list.append(f"Row {row_num}: {str(e)}")
    
    elif import_type == 'sales':
        for index, row in df.iterrows():
            row_num = index + 2
            row_data = row.to_dict()
            try:
                product_code = str(row.get('product_code', '')).strip()
                product = Product.objects.get(code=product_code)
                qty = int(row.get('quantity', 0))
                if qty <= 0:
                    success_count += 1
                    continue
                order = Order.objects.create(
                    order_type='walk_in',
                    status='confirmed',
                    payment_status='paid',
                    subtotal=product.price * qty,
                    tax=0,
                    total=product.price * qty,
                    created_by=request.user
                )
                OrderItem.objects.create(order=order, product=product, quantity=qty, price=product.price, subtotal=product.price * qty)
                old_stock = product.stock
                product.stock -= qty
                product.save()
                InventoryTransaction.objects.create(
                    product=product, transaction_type='out', quantity=qty,
                    previous_stock=old_stock, new_stock=product.stock,
                    reference=order.order_number, notes="Historical sales import",
                    created_by=request.user
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                error_rows.append({'row': row_num, 'data': row_data, 'error': str(e)})
                errors_list.append(f"Row {row_num}: {str(e)}")
    
    return success_count, error_count, error_rows


def _create_error_report_file(error_rows, import_type, file_name):
    """Create Excel error report and return HttpResponse or None"""
    if not error_rows:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Errors"
    ws.cell(row=1, column=1, value="Row #")
    ws.cell(row=1, column=2, value="Error Message")
    for c, col in enumerate(error_rows[0].get('data', {}).keys(), 3):
        ws.cell(row=1, column=c, value=str(col))
    for idx, er in enumerate(error_rows[:500], 2):
        ws.cell(row=idx, column=1, value=er['row'])
        ws.cell(row=idx, column=2, value=er['error'])
        for c, (k, v) in enumerate(er.get('data', {}).items(), 3):
            ws.cell(row=idx, column=c, value=str(v) if pd.notna(v) else '')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    base = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
    response['Content-Disposition'] = f'attachment; filename="import_errors_{import_type}_{base}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin_or_manager)
def import_data(request):
    if request.method == 'POST':
        import_type = request.POST.get('import_type')
        files = request.FILES.getlist('files') or request.FILES.getlist('file')
        if not files:
            single = request.FILES.get('file')
            if single:
                files = [single]
        if not files or not import_type:
            messages.error(request, 'Please select import type and at least one file.')
            return redirect('import_data')
        
        all_errors = []
        total_success = 0
        total_failed = 0
        all_error_rows = []
        file_names = []
        
        with transaction.atomic():
            for file in files:
                if not file.name:
                    continue
                try:
                    if file.name.lower().endswith('.csv'):
                        df = pd.read_csv(file)
                    elif file.name.lower().endswith(('.xlsx', '.xls')):
                        df = pd.read_excel(file)
                    else:
                        messages.warning(request, f'Skipped {file.name}: unsupported format.')
                        continue
                except Exception as e:
                    messages.error(request, f'Could not read {file.name}: {e}')
                    continue
                
                file_names.append(file.name)
                success, failed, error_rows = _process_import_rows(import_type, df, file.name, request, all_errors)
                total_success += success
                total_failed += failed
                for er in error_rows:
                    er['file'] = file.name
                    all_error_rows.append(er)
            
            # Create import history record
            ImportHistory.objects.create(
                import_type=import_type,
                file_name=', '.join(file_names),
                status='success' if total_failed == 0 else 'partial' if total_success > 0 else 'failed',
                total_records=total_success + total_failed,
                success_records=total_success,
                failed_records=total_failed,
                error_log='\n'.join(all_errors[:200]),
                imported_by=request.user
            )
        
        if total_failed > 0 and all_error_rows:
            err_resp = _create_error_report_file(all_error_rows, import_type, file_names[0] if file_names else 'import')
            if err_resp:
                messages.warning(request, f'Import completed: {total_success} success, {total_failed} errors. Download error report below.')
                request.session['import_error_report'] = None
                return err_resp
        if total_success > 0:
            messages.success(request, f'Successfully imported {total_success} records from {len(file_names)} file(s).')
        elif total_failed > 0:
            messages.error(request, f'Import failed. {total_failed} errors. Check file format and data.')
        
        return redirect('import_history')
    
    return render(request, 'core/import_data.html')

@login_required
def import_history(request):
    imports = ImportHistory.objects.all().order_by('-created_at')
    return render(request, 'core/import_history.html', {'imports': imports})


@login_required
@user_passes_test(is_admin_or_manager)
def download_import_template(request, import_type):
    """Download CSV template for import."""
    templates = {
        'product': (['code', 'name', 'category', 'price', 'cost', 'stock', 'description'], 'products_template.csv'),
        'customer': (['name', 'email', 'phone', 'address'], 'customers_template.csv'),
        'inventory': (['product_code', 'new_stock'], 'inventory_template.csv'),
        'sales': (['product_code', 'quantity'], 'sales_template.csv'),
    }
    if import_type not in templates:
        return redirect('import_data')
    headers, filename = templates[import_type]
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # BOM for Excel UTF-8
    writer = csv.writer(response)
    writer.writerow(headers)
    if import_type == 'product':
        writer.writerow(['BREAD001', 'Sample Bread', 'bread', '50.00', '25.00', '100', 'Description'])
    return response


@login_required
def export_data(request):
    export_type = request.GET.get('type', 'products')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    
    if export_type == 'products':
        ws.title = "Products"
        headers = ['Code', 'Name', 'Category', 'Price', 'Cost', 'Stock', 'Low Stock Threshold', 'Created At']
        data = Product.objects.all().values_list(
            'code', 'name', 'category', 'price', 'cost', 'stock', 'low_stock_threshold', 'created_at'
        )
    
    elif export_type == 'customers':
        ws.title = "Customers"
        headers = ['Name', 'Email', 'Phone', 'Address', 'Loyalty Points', 'Created At']
        data = Customer.objects.all().values_list(
            'name', 'email', 'phone', 'address', 'loyalty_points', 'created_at'
        )
    
    elif export_type == 'orders':
        ws.title = "Orders"
        headers = ['Order Number', 'Customer', 'Date', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total', 'Status']
        data = []
        for order in Order.objects.all().select_related('customer'):
            data.append((
                order.order_number,
                order.customer.name if order.customer else 'Walk-in',
                order.created_at.strftime('%Y-%m-%d %H:%M'),
                order.items.count(),
                order.subtotal,
                order.tax,
                order.discount,
                order.total,
                order.status
            ))
    
    elif export_type == 'inventory':
        ws.title = "Inventory"
        headers = ['Product', 'Current Stock', 'Low Stock Threshold', 'Status']
        data = Product.objects.all().values_list('name', 'stock', 'low_stock_threshold')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{export_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

# ========== PUBLIC VIEWS ==========

def privacy_policy(request):
    from django.http import HttpResponse
    html = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Pandeli Privacy Policy</title><style>:root{--primary:#D39A73;--accent:#B8744C;--text:#3E2C23;--bg:#fdf8f5;--card:#ffffff;}body{margin:0;font-family:'Segoe UI',sans-serif;background:var(--bg);color:var(--text);}.header{background:linear-gradient(135deg,var(--primary),var(--accent));color:white;padding:40px 20px;text-align:center;border-bottom-left-radius:30px;border-bottom-right-radius:30px;}.header h1{margin:0;}.container{max-width:900px;margin:-20px auto 40px;padding:20px;}.card{background:var(--card);padding:20px;margin-bottom:20px;border-radius:15px;box-shadow:0 4px 12px rgba(0,0,0,0.05);}h2{color:var(--accent);}.notice{background:rgba(211,154,115,0.15);border-left:4px solid var(--primary);padding:12px;border-radius:8px;margin-top:10px;}.footer{text-align:center;font-size:13px;color:#888;padding:20px;}</style></head><body><div class="header"><h1>Pandeli Privacy Policy</h1><p>Last updated: April 16, 2026</p></div><div class="container"><div class="card"><p>This Privacy Policy explains how we collect, use, and protect your data.</p><div class="notice">By using the Service, you agree to this Privacy Policy.</div></div><div class="card"><h2>Data We Collect</h2><ul><li>Email, name, phone number</li><li>Usage data (IP, device, activity)</li><li>Location, contacts, camera (with permission)</li></ul></div><div class="card"><h2>How We Use Data</h2><ul><li>Provide and maintain service</li><li>Improve user experience</li><li>Send updates and notifications</li></ul></div><div class="card"><h2>Your Rights</h2><ul><li>Access your data</li><li>Edit or delete your information</li></ul></div><div class="card"><h2>Contact</h2><p>Email: <strong>pandelibakehouse@gmail.com</strong></p></div></div><div class="footer">&copy; 2026 Pandeli</div></body></html>"""
    return HttpResponse(html)

# ========== APP FEATURES VIEWS ==========

@login_required
def app_feature_list(request):
    try:
        features = list(AppFeature.objects.all())
    except Exception:
        features = []
    try:
        bundles = list(Bundle.objects.all())
    except Exception:
        bundles = []
    return render(request, 'core/app_feature_list.html', {'features': features, 'bundles': bundles})

@login_required
@require_POST
def app_feature_add(request):
    title = request.POST.get('title', '')
    subtitle = request.POST.get('subtitle', '')
    order = int(request.POST.get('order', 0))
    image = request.FILES.get('image')
    if not image:
        messages.error(request, 'Image is required.')
        return redirect('app_feature_list')
    try:
        AppFeature.objects.create(title=title, subtitle=subtitle, image=image, order=order)
        messages.success(request, 'App feature added successfully.')
    except Exception as e:
        messages.error(request, f'Error uploading image: {str(e)}')
    return redirect('app_feature_list')

@login_required
@require_POST
def app_feature_toggle(request, pk):
    feature = get_object_or_404(AppFeature, pk=pk)
    feature.is_active = not feature.is_active
    feature.save(update_fields=['is_active'])
    return JsonResponse({'is_active': feature.is_active})

@login_required
@require_POST
def app_feature_delete(request, pk):
    feature = get_object_or_404(AppFeature, pk=pk)
    try:
        if feature.image:
            feature.image.delete(save=False)
    except Exception:
        pass
    feature.delete()
    messages.success(request, 'App feature deleted.')
    return redirect('app_feature_list')

def debug_order_request(request):
    """Debug endpoint — logs what the app sends when creating an order."""
    import logging
    logger = logging.getLogger(__name__)
    try:
        body = request.body.decode('utf-8')
    except Exception:
        body = str(request.body)
    logger.info(f"DEBUG ORDER REQUEST - Method: {request.method}")
    logger.info(f"DEBUG ORDER REQUEST - Headers: {dict(request.headers)}")
    logger.info(f"DEBUG ORDER REQUEST - Body: {body}")
    logger.info(f"DEBUG ORDER REQUEST - POST: {dict(request.POST)}")
    return JsonResponse({
        'method': request.method,
        'body': body,
        'post': dict(request.POST),
        'content_type': request.content_type,
    })

# ========== BUNDLE VIEWS ==========

@login_required
def bundle_list(request):
    return redirect('app_feature_list')

@login_required
@require_POST
def bundle_add(request):
    name = request.POST.get('name', '')
    description = request.POST.get('description', '')
    subtitle = request.POST.get('subtitle', '')
    item_count = int(request.POST.get('item_count', 1))
    # Handle multiple categories (checkboxes)
    categories = request.POST.getlist('categories')
    category = ','.join(categories) if categories else ''
    order = int(request.POST.get('order', 0))
    image = request.FILES.get('image')
    try:
        Bundle.objects.create(
            name=name, description=description,
            subtitle=subtitle, item_count=item_count,
            category=category, order=order,
            image=image
        )
        messages.success(request, 'Bundle added successfully.')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
    return redirect('app_feature_list')

@login_required
@require_POST
def bundle_toggle(request, pk):
    bundle = get_object_or_404(Bundle, pk=pk)
    bundle.is_active = not bundle.is_active
    bundle.save(update_fields=['is_active'])
    return JsonResponse({'is_active': bundle.is_active})

@login_required
@require_POST
def bundle_edit(request, pk):
    bundle = get_object_or_404(Bundle, pk=pk)
    bundle.name = request.POST.get('name', bundle.name)
    bundle.subtitle = request.POST.get('subtitle', bundle.subtitle)
    bundle.description = request.POST.get('description', bundle.description)
    bundle.item_count = int(request.POST.get('item_count', bundle.item_count))
    categories = request.POST.getlist('categories')
    bundle.category = ','.join(categories) if categories else ''
    bundle.order = int(request.POST.get('order', bundle.order))
    image = request.FILES.get('image')
    if image:
        try:
            if bundle.image:
                bundle.image.delete(save=False)
        except Exception:
            pass
        bundle.image = image
    bundle.save()
    messages.success(request, f'Bundle "{bundle.name}" updated.')
    return redirect('app_feature_list')

@login_required
@require_POST
def bundle_delete(request, pk):
    bundle = get_object_or_404(Bundle, pk=pk)
    try:
        if bundle.image:
            bundle.image.delete(save=False)
    except Exception:
        pass
    bundle.delete()
    messages.success(request, 'Bundle deleted.')
    return redirect('app_feature_list')
    bundle = get_object_or_404(Bundle, pk=pk)
    try:
        if bundle.image:
            bundle.image.delete(save=False)
    except Exception:
        pass
    bundle.delete()
    messages.success(request, 'Bundle deleted.')
    return redirect('app_feature_list')

def bundles_api(request):
    """Public API — returns active bundles with filtered products by category."""
    try:
        bundles = list(Bundle.objects.filter(is_active=True).order_by('order', '-created_at'))
    except Exception:
        return JsonResponse([], safe=False)

    if not bundles:
        return JsonResponse([], safe=False)

    # Fetch ALL available products once — avoid N+1 query per bundle
    raw_products = list(
        Product.objects.filter(is_archived=False, is_available=True)
        .values('id', 'name', 'price', 'category', 'description', 'image')
    )

    # Build enriched product dicts grouped by category
    products_by_category = {}
    all_products_list = []
    for p in raw_products:
        img = p.get('image')
        if img:
            try:
                from django.conf import settings
                image_url = request.build_absolute_uri(settings.MEDIA_URL + img)
            except Exception:
                image_url = None
        else:
            image_url = None

        enriched = {
            'id': p['id'],
            'name': p['name'],
            'price': float(p['price']),
            'category': p['category'],
            'description': p['description'],
            'image_url': image_url,
        }
        cat = p['category']
        if cat not in products_by_category:
            products_by_category[cat] = []
        products_by_category[cat].append(enriched)
        all_products_list.append(enriched)

    data = []
    for b in bundles:
        # Filter products by bundle category
        if b.category:
            cats = [c.strip() for c in b.category.split(',') if c.strip()]
            bundle_products = []
            for cat in cats:
                bundle_products.extend(products_by_category.get(cat, []))
        else:
            bundle_products = all_products_list

        image_url = None
        if b.image:
            try:
                image_url = request.build_absolute_uri(b.image.url)
            except Exception:
                image_url = b.image.url

        data.append({
            'id': b.id,
            'name': b.name,
            'description': b.description,
            'subtitle': b.subtitle,
            'item_count': b.item_count,
            'category': b.category,
            'image_url': image_url,
            'order': b.order,
            'products': bundle_products,
        })
    return JsonResponse(data, safe=False)

@csrf_exempt
def order_webhook(request):
    """
    Webhook endpoint called by Supabase when a new order is created.
    Creates a notification for all staff/admin users.
    Must be csrf_exempt because Supabase does not send a CSRF token.
    """
    import json as _json
    import logging
    _logger = logging.getLogger(__name__)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    try:
        body = _json.loads(request.body)
    except (_json.JSONDecodeError, Exception) as e:
        _logger.error(f"order_webhook: invalid JSON body — {e}")
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    try:
        record = body.get('record', {})
        order_number = record.get('order_number') or body.get('order_number', 'Unknown')
        total = record.get('total') or body.get('total', 0)
        order_type = record.get('order_type') or body.get('order_type', 'online')

        from .notifications import NotificationService
        NotificationService.notify_admins(
            title=f"🛒 New Order #{order_number}",
            message=f"New {order_type} order received. Total: ₱{total}",
            notification_type='order',
            priority='high',
            link='/orders/',
            action_text='View Orders',
        )

        _logger.info(f"order_webhook: notifications sent for order #{order_number}")
        return JsonResponse({'status': 'ok'})

    except Exception as e:
        _logger.error(f"order_webhook error: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def app_features_api(request):
    features = AppFeature.objects.filter(is_active=True).order_by('order', '-created_at')
    data = []
    for f in features:
        image_url = None
        if f.image:
            try:
                image_url = request.build_absolute_uri(f.image.url)
            except Exception:
                image_url = f.image.url
        data.append({'id': f.id, 'title': f.title, 'subtitle': f.subtitle, 'image_url': image_url, 'order': f.order})
    return JsonResponse(data, safe=False)

# ========== SUPPLIER VIEWS ==========

@login_required
def supplier_list(request):
    suppliers = Supplier.objects.all()
    
    # Search
    search = request.GET.get('search')
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) | 
            Q(email__icontains=search) |
            Q(phone__icontains=search)
        )
    
    context = {
        'suppliers': suppliers,
    }
    return render(request, 'core/supplier_list.html', context)

# ========== MESSAGES/NOTIFICATIONS VIEWS ==========

@login_required
def messages_view(request):
    notifications = Notification.objects.filter(
        recipient_user=request.user
    ).order_by('-created_at')
    
    # Filter by type
    notification_type = request.GET.get('type', 'all')
    if notification_type == 'unread':
        notifications = notifications.filter(is_read=False)
    elif notification_type not in ('all', ''):
        notifications = notifications.filter(notification_type=notification_type)
    
    # Mark as read
    notification_id = request.GET.get('read')
    if notification_id:
        try:
            notification = Notification.objects.get(
                id=notification_id,
                recipient_user=request.user
            )
            notification.is_read = True
            notification.save()
        except Notification.DoesNotExist:
            pass
    
    # Get unread count
    unread_count = Notification.objects.filter(
        recipient_user=request.user,
        is_read=False
    ).count()
    
    context = {
        'notifications': notifications,
        'unread_count': unread_count,
    }
    return render(request, 'core/messages.html', context)

# ========== POINT OF SALE (POS) VIEWS ==========

@login_required
def pos_view(request):
    """
    Point of Sale interface for walk-in customers
    """
    # Get all available products
    # Show all products in POS so the cashier can see items even when temporarily out of stock.
    # Ordering/checkout is still blocked server-side by `pos_create_order` when stock is insufficient.
    products = Product.objects.filter(is_archived=False).order_by('category', 'name')
    
    # Get categories for filtering
    categories = Product.CATEGORY_CHOICES
    
    # Get recent orders (last 10)
    recent_orders = Order.objects.filter(
        order_type='walk_in'
    ).select_related('customer').order_by('-created_at')[:10]
    
    # Get customers for quick selection
    customers = Customer.objects.all().order_by('name')[:20]
    
    context = {
        'products': products,
        'categories': categories,
        'recent_orders': recent_orders,
        'customers': customers,
    }
    return render(request, 'core/pos.html', context)


@login_required
@transaction.atomic
def pos_create_order(request):
    """
    AJAX view to create a new POS order
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            customer_id = data.get('customer_id')
            payment_method = (data.get('payment_method') or 'cash').lower()
            discount = data.get('discount', 0)
            amount_received = data.get('amount_received', 0)
            notes = data.get('notes', '')
            
            if not items:
                return JsonResponse({'error': 'No items in cart'}, status=400)

            if payment_method not in ('cash', 'gcash'):
                return JsonResponse({'error': 'Invalid payment method'}, status=400)
            
            # Calculate totals
            subtotal = 0
            order_items_data = []
            
            for item in items:
                product = Product.objects.get(id=item['product_id'])
                quantity = int(item['quantity'])
                
                if product.stock < quantity:
                    return JsonResponse({
                        'error': f'Insufficient stock for {product.name}. Available: {product.stock}'
                    }, status=400)
                
                item_subtotal = product.price * quantity
                subtotal += item_subtotal
                
                order_items_data.append({
                    'product': product,
                    'quantity': quantity,
                    'price': product.price,
                    'subtotal': item_subtotal
                })

            # Validate required raw materials for all items before creating/updating anything.
            material_requirements = {}
            for item_data in order_items_data:
                product = item_data['product']
                quantity = item_data['quantity']
                for recipe in product.recipe.select_related('raw_material').all():
                    material_id = recipe.raw_material_id
                    needed_qty = float(recipe.quantity) * quantity
                    if material_id not in material_requirements:
                        material_requirements[material_id] = {
                            'material': recipe.raw_material,
                            'needed': 0.0,
                        }
                    material_requirements[material_id]['needed'] += needed_qty

            insufficient_materials = []
            for data_item in material_requirements.values():
                material = data_item['material']
                needed = data_item['needed']
                if float(material.stock_quantity) < needed:
                    insufficient_materials.append(
                        f"{material.name} (available: {material.stock_quantity}, needed: {needed:.2f})"
                    )

            if insufficient_materials:
                return JsonResponse({
                    'error': f"Insufficient raw materials: {', '.join(insufficient_materials)}"
                }, status=400)
            
            # Discount only (tax removed from POS)
            try:
                discount = float(discount or 0)
            except (TypeError, ValueError):
                return JsonResponse({'error': 'Invalid discount'}, status=400)

            if discount < 0:
                return JsonResponse({'error': 'Discount cannot be negative'}, status=400)

            tax = 0
            total = float(subtotal) - discount

            if total < 0:
                return JsonResponse({'error': 'Total cannot be negative'}, status=400)

            # Payment validation
            try:
                amount_received = float(amount_received or 0)
            except (TypeError, ValueError):
                return JsonResponse({'error': 'Invalid amount received'}, status=400)

            if payment_method == 'cash':
                if amount_received < total:
                    return JsonResponse({'error': 'Insufficient cash amount received'}, status=400)
                change_amount = amount_received - total
            else:
                # Gcash: usually exact amount, but allow >= total; no change returned
                if amount_received and amount_received < total:
                    return JsonResponse({'error': 'Insufficient Gcash amount'}, status=400)
                amount_received = total
                change_amount = 0
            
            # Create order
            customer = None
            if customer_id:
                customer = Customer.objects.get(id=customer_id)
            
            order = Order.objects.create(
                customer=customer,
                order_type='walk_in',
                status='confirmed',  # POS orders are auto-confirmed
                payment_status='paid',
                payment_method=payment_method,
                amount_received=amount_received,
                change_amount=change_amount,
                subtotal=subtotal,
                tax=tax,
                discount=discount,
                total=total,
                notes=notes,
                created_by=request.user
            )
            
            # Create order items and update stock
            for item_data in order_items_data:
                product = item_data['product']
                quantity = item_data['quantity']
                
                # Create order item
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    price=product.price,
                    subtotal=product.price * quantity
                )
                
                # Update stock with row-level lock to prevent race conditions
                product = Product.objects.select_for_update().get(pk=product.pk)
                old_stock = product.stock
                product.stock -= quantity
                product.save()

                # Record inventory transaction
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='out',
                    quantity=quantity,
                    previous_stock=old_stock,
                    new_stock=product.stock,
                    reference=order.order_number,
                    notes=f"POS sale",
                    created_by=request.user
                )

                # Notify if low stock after deduction
                from .notifications import notify_if_low_stock, notify_if_material_low
                notify_if_low_stock(product)

                # Deduct raw materials based on recipe
                for recipe in product.recipe.select_related('raw_material').all():
                    material = recipe.raw_material
                    needed_qty = float(recipe.quantity) * quantity
                    old_material_stock = float(material.stock_quantity)
                    material.stock_quantity = old_material_stock - needed_qty
                    material.save()

                    RawMaterialTransaction.objects.create(
                        raw_material=material,
                        transaction_type='out',
                        quantity=needed_qty,
                        previous_stock=old_material_stock,
                        new_stock=material.stock_quantity,
                        reference=order.order_number,
                        notes=f"Used for POS sale: {product.name}",
                        created_by=request.user
                    )
                    notify_if_material_low(material)
            
            # Update customer loyalty points (optional)
            if customer:
                customer.loyalty_points += int(total / 10)  # 1 point per $10 spent
                customer.save()
            
            return JsonResponse({
                'success': True,
                'order_id': order.id,
                'order_number': order.order_number,
                'total': float(total),
                'change_amount': float(change_amount),
            })
            
        except Product.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=400)
        except Customer.DoesNotExist:
            return JsonResponse({'error': 'Customer not found'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def pos_get_product(request, product_id):
    """
    AJAX view to get product details
    """
    try:
        product = Product.objects.get(id=product_id, is_available=True)
        data = {
            'id': product.id,
            'name': product.name,
            'price': float(product.price),
            'stock': product.stock,
            'code': product.code,
            'category': product.get_category_display(),
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)


@login_required_json
def order_modal_products(request):
    """
    AJAX: return all products for the Create New Order modal (no is_available filter).
    """
    try:
        products = Product.objects.all().order_by('name')[:200]
        data = []
        for p in products:
            data.append({
                'id': p.id,
                'name': p.name or '',
                'price': float(p.price) if p.price is not None else 0,
                'stock': int(p.stock) if p.stock is not None else 0,
                'code': getattr(p, 'code', '') or '',
            })
        return JsonResponse({'products': data})
    except Exception as e:
        return JsonResponse({'error': str(e), 'products': []}, status=500)


@login_required
def pos_search_products(request):
    """
    AJAX view to search products
    """
    query = request.GET.get('q', '')
    category = request.GET.get('category', '')
    
    products = Product.objects.filter(is_available=True, stock__gt=0)
    
    if query:
        products = products.filter(
            Q(name__icontains=query) | 
            Q(code__icontains=query) |
            Q(description__icontains=query)
        )
    
    if category:
        products = products.filter(category=category)
    
    products = products.order_by('name')[:50]
    
    data = [{
        'id': p.id,
        'name': p.name,
        'price': float(p.price),
        'stock': p.stock,
        'code': p.code,
        'category': p.get_category_display(),
    } for p in products]
    
    return JsonResponse({'products': data})


@login_required
def pos_receipt(request, order_id):
    """
    View to display/print receipt
    """
    order = get_object_or_404(Order, id=order_id)

    # Safely resolve customer name — handles UUID customer_id from mobile app
    customer_name = 'Walk-in Customer'
    try:
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT customer_id FROM core_order WHERE id = %s", [order_id])
            row = cursor.fetchone()
            if row and row[0]:
                try:
                    cid = int(row[0])
                    from .models import Customer
                    customer = Customer.objects.filter(pk=cid).first()
                    if customer:
                        customer_name = customer.name
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass

    # Get return URL from query parameter (default to POS page)
    return_url = request.GET.get('return_url', 'pos')
    
    return render(request, 'core/pos_receipt.html', {
        'order': order,
        'customer_name': customer_name,
        'return_url': return_url,
    })

@login_required
def notifications_api(request):
    """
    API endpoint for notifications (AJAX)
    """
    notification_type = request.GET.get('type', 'all')
    limit = int(request.GET.get('limit', 20))
    
    notifications = Notification.objects.filter(
        recipient_user=request.user
    )

    if notification_type == 'unread':
        notifications = notifications.filter(is_read=False)
    elif notification_type != 'all':
        notifications = notifications.filter(notification_type=notification_type)
    
    notifications = notifications.order_by('-created_at')[:limit]
    
    data = [{
        'id': n.id,
        'title': n.title,
        'message': n.message[:100] + ('...' if len(n.message) > 100 else ''),
        'type': n.notification_type,
        'priority': n.priority,
        'is_read': n.is_read,
        'link': n.link,
        'action_text': n.action_text,
        'created_at': n.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        'time_ago': timesince(n.created_at),
    } for n in notifications]
    
    return JsonResponse({'notifications': data})


@login_required
@require_POST
def mark_notification_read(request, notification_id):
    """Mark a single notification as read"""
    try:
        notification = Notification.objects.get(
            recipient_user=request.user,
            id=notification_id
        )
        notification.mark_as_read()
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)


@login_required
@require_POST
def mark_all_notifications_read(request):
    """Mark all notifications as read for the current user"""
    count = Notification.objects.filter(
        recipient_user=request.user,
        is_read=False
    ).update(is_read=True)
    return JsonResponse({'success': True, 'count': count})


@login_required
def notification_count(request):
    """Get unread notification count — also catches any new orders that slipped through"""
    # Check for orders that have no notification yet (e.g. inserted directly via Supabase)
    try:
        _create_missing_order_notifications()
    except Exception:
        pass

    count = Notification.objects.filter(
        recipient_user=request.user,
        is_read=False
    ).count()
    return JsonResponse({'count': count})


def _create_missing_order_notifications():
    """
    Find orders that have no 'new order' notification and create one per staff user.
    Uses get_or_create so it is safe to call repeatedly — no duplicates.
    """
    from django.utils import timezone
    from datetime import timedelta
    from django.contrib.auth.models import User as AuthUser
    from .models import UserProfile

    cutoff = timezone.now() - timedelta(hours=24)
    new_orders = Order.objects.filter(created_at__gte=cutoff)
    if not new_orders.exists():
        return

    # Collect all staff user IDs (same logic as notify_admins)
    user_ids = set(AuthUser.objects.filter(is_superuser=True, is_active=True).values_list('id', flat=True))
    user_ids |= set(UserProfile.objects.filter(
        role__in=['admin', 'production_admin', 'manager', 'cashier', 'staff']
    ).values_list('user_id', flat=True))
    staff_users = list(AuthUser.objects.filter(id__in=user_ids, is_active=True))

    for order in new_orders:
        title = f"🛒 New Order #{order.order_number}"
        link = f'/orders/{order.pk}/'
        message = f"New {order.get_order_type_display()} order received. Total: ₱{order.total}"
        for user in staff_users:
            Notification.objects.get_or_create(
                notification_type='order',
                recipient_user=user,
                link=link,
                title=title,
                defaults={
                    'message': message,
                    'recipient_type': 'admin',
                    'priority': 'high',
                    'is_read': False,
                    'action_text': 'View Order',
                }
            )


@login_required
@user_passes_test(is_admin_or_manager)
def trigger_low_stock_check(request):
    """Manually trigger low stock check (admin only)"""
    if request.method == 'POST':
        count = check_low_stock_and_notify()
        messages.success(request, f'Low stock check completed. {count} notifications created.')
        return redirect('inventory_status')
    return redirect('inventory_status')


@login_required
@user_passes_test(is_admin_or_manager)
def create_bulk_notification(request):
    """Create a notification for multiple users (admin only)"""
    if request.method == 'POST':
        title = request.POST.get('title')
        message = request.POST.get('message')
        notification_type = request.POST.get('notification_type', 'system')
        recipient_type = request.POST.get('recipient_type', 'admin')
        
        if recipient_type == 'all_admins':
            NotificationService.notify_admins(title, message, notification_type)
            messages.success(request, f'Notification sent to all admins')
        elif recipient_type == 'all_staff':
            NotificationService.notify_staff(title, message, notification_type)
            messages.success(request, f'Notification sent to all staff')
        
        return redirect('messages')
    
    return render(request, 'core/create_notification.html')

def index(request):
    """Simple index view that redirects to login"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


# ========== USER MANAGEMENT VIEWS (admin/manager only) ==========

def _require_admin(request):
    """Return True if the user is allowed to manage other users."""
    if request.user.is_superuser:
        return True
    try:
        return request.user.profile.role in ('admin', 'manager')
    except Exception:
        return False


@login_required
def user_list(request):
    if not _require_admin(request):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    from django.contrib.auth.models import User as AuthUser
    users = AuthUser.objects.select_related('profile').order_by('username')

    # Search
    q = request.GET.get('q', '').strip()
    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q)
        )

    # Role filter
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(profile__role=role_filter)

    return render(request, 'core/user_list.html', {
        'users': users,
        'role_choices': UserProfile.ROLE_CHOICES,
        'q': q,
        'role_filter': role_filter,
        'role_matrix': [
            # (module, [admin, manager, cashier, staff, production_admin])
            ('Dashboard',        [True,  True,  False, False, False]),
            ('Orders',           [True,  True,  True,  True,  True ]),
            ('POS',              [True,  True,  True,  False, False]),
            ('Products',         [True,  True,  True,  False, True ]),
            ('Inventory',        [True,  True,  False, False, True ]),
            ('Suppliers',        [True,  True,  False, False, False]),
            ('Sales Report',     [True,  True,  False, False, False]),
            ('Cost Analytics',   [True,  True,  False, False, False]),
            ('Forecast',         [True,  True,  False, False, False]),
            ('Import / Export',  [True,  True,  False, False, False]),
            ('User Management',  [True,  False, False, False, False]),
            ('Notifications',    [True,  True,  True,  True,  True ]),
        ],
    })


@login_required
def user_create(request):
    if not _require_admin(request):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    from .forms import UserCreateForm
    from django.contrib.auth.models import User as AuthUser

    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = form.cleaned_data['role']
            profile.phone = form.cleaned_data.get('phone', '')
            profile.save()
            messages.success(request, f'User "{user.username}" created successfully.')
            return redirect('user_list')
    else:
        form = UserCreateForm()

    return render(request, 'core/user_form.html', {'form': form, 'action': 'Create'})


@login_required
def user_edit(request, pk):
    if not _require_admin(request):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    from django.contrib.auth.models import User as AuthUser
    from .forms import UserEditForm

    target_user = get_object_or_404(AuthUser, pk=pk)

    # Prevent non-superusers from editing superusers
    if target_user.is_superuser and not request.user.is_superuser:
        messages.error(request, 'You cannot edit a superuser account.')
        return redirect('user_list')

    profile, _ = UserProfile.objects.get_or_create(user=target_user)

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=target_user)
        if form.is_valid():
            user = form.save(commit=False)
            new_pw = form.cleaned_data.get('new_password')
            if new_pw:
                user.set_password(new_pw)
            user.save()
            profile.role = form.cleaned_data['role']
            profile.phone = form.cleaned_data.get('phone', '')
            profile.save()
            messages.success(request, f'User "{user.username}" updated successfully.')
            return redirect('user_list')
    else:
        form = UserEditForm(instance=target_user, initial={
            'role': profile.role,
            'phone': profile.phone,
        })

    return render(request, 'core/user_form.html', {
        'form': form,
        'action': 'Edit',
        'target_user': target_user,
        'profile': profile,
    })


@login_required
@require_POST
def user_delete(request, pk):
    if not _require_admin(request):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    from django.contrib.auth.models import User as AuthUser
    target_user = get_object_or_404(AuthUser, pk=pk)

    # Cannot delete yourself or a superuser (unless you are one)
    if target_user == request.user:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('user_list')
    if target_user.is_superuser and not request.user.is_superuser:
        messages.error(request, 'You cannot delete a superuser account.')
        return redirect('user_list')

    username = target_user.username
    target_user.delete()
    messages.success(request, f'User "{username}" deleted.')
    return redirect('user_list')


@login_required
@require_POST
def user_toggle_active(request, pk):
    if not _require_admin(request):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    from django.contrib.auth.models import User as AuthUser
    target_user = get_object_or_404(AuthUser, pk=pk)

    if target_user == request.user:
        messages.error(request, 'You cannot deactivate your own account.')
        return redirect('user_list')

    target_user.is_active = not target_user.is_active
    target_user.save(update_fields=['is_active'])
    state = 'activated' if target_user.is_active else 'deactivated'
    messages.success(request, f'User "{target_user.username}" {state}.')
    return redirect('user_list')
